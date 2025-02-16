from openpyxl import load_workbook 
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

month = 'February'

data_dir = 'D:/Inna_home/Study/Projects/Excel_auto_reports/'
wb = load_workbook(data_dir+'pivot_table.xlsx')
sheet = wb['Report']

#space
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

#Barchart
barchart = BarChart ()

#Locate data and categories of the barchart
data = Reference (sheet, min_col=min_column+1, max_col=max_column+1, min_row=min_row, max_row=max_row)
categories = Reference (sheet, min_col=min_column, max_col=max_column, min_row=min_row+1, max_row=max_row)

#Adding data and categories
barchart.add_data(data, titles_from_data=True)
barchart.set_categories (categories)

#Make chart
sheet.add_chart (barchart, 'B12')
barchart.title = 'Sales by Product line'
barchart.style = 5 #style chart bar

#Loop throught columns with data
for i in range (min_column+1, max_column+1):
    #print(i)
    letter= (get_column_letter(i))
    sheet [f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet [f'{letter}{max_row+1}'].style = 'Currency'

sheet['A1'] = 'Sales report'
sheet['A2'] = month
sheet['A1'].font = Font ('Arial', bold=True, size= 20)
sheet['A2'].font = Font ('Arial', bold= True, size = 10)

wb.save(data_dir+f'report_{month}.xlsx')

