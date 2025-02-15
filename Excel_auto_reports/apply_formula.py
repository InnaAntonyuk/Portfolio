from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

data_dir = 'D:/Inna_home/Study/Projects/Excel_auto_reports/'
wb = load_workbook(data_dir + 'barchart.xlsx')
sheet = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

#sheet['B8'] = '=Sum(B6:B7)'
#sheet['B8'].style = 'Currency'

for i in range (min_column+1, max_column+1):
    #print(i)
    letter= (get_column_letter(i))
    sheet [f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet [f'{letter}{max_row+1}'].style = 'Currency'
wb.save(data_dir + 'report.xlsx')